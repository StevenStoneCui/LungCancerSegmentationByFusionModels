# -*- coding: utf-8 -*-
"""
Created on Mon Apr  5 14:36:29 2021

@author: ariken
"""

import tensorflow as tf
from tensorflow.keras.layers import Input, Dropout, Conv3D, Conv3DTranspose
from tensorflow.keras.layers import BatchNormalization, Activation
from tensorflow.keras.layers import concatenate
from tensorflow.keras.models import Model
from RandomFlipLayer import RandomFlipLayer
from RandomRotationLayer import RandomRotationLayer
from RandomSpatialScalingLayer import RandomSpatialScalingLayer
from tensorflow.keras.optimizers import RMSprop
import tensorflow.keras.backend as K
import Vnet
import os 
import sys
import random
import cupy as cp
import cupyx.scipy.ndimage as cpndi
import copy
import numpy as np
import scipy.io as scio
import warnings
from PIL import Image, ImageDraw
import openpyxl
import gc

warnings.filterwarnings('ignore', '.*output shape of zoom.*')

os.environ["CUDA_VISIBLE_DEVICES"] = "0"
globalVar = []

config = tf.compat.v1.ConfigProto()  
config.gpu_options.allow_growth=True   
session = tf.compat.v1.Session(config=config)
tf.compat.v1.keras.backend.set_session(session)
#tf.config.run_functions_eagerly(True)
mempool = cp.get_default_memory_pool()
pinned_mempool = cp.get_default_pinned_memory_pool()
sys.setrecursionlimit(1000000000)
class Stack(object):
    def __init__(self, length = 1):
        self.items = []
        self.len = length

    def is_empty(self):
        return self.items == []

    def peek(self):
        return self.items[len(self.items) - 1]

    def size(self):
        return len(self.items)

    def push(self, item):
        if len(self.items) <= self.len:
            self.items.append(item)
        else:
            raise Exception("Stack overflow")

    def is_full(self):
        return self.len == len(self.items)

    def pop(self):
        if self.is_empty() is False:
            return self.items.pop()
        else:
            raise Exception("Stack empty")
     
    def popMore(self, num = 1):
        if self.size() < num:
            return [] 
        result = []
        for _ in range(num):
            result.append(self.pop())
        random.shuffle(result)
        return result        
        
    def clear(self):
        self.items.clear()
    
    def shuffle(self):
        random.shuffle(self.items)

def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet()
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=value[i][j])
    workbook.save(path)

def resizing(img, label, windowsSize = (40, 40, 40), interpOrderImg = 3, interpOrderLabel = 0):
    imgResized = []
    labelResized = []
    for x in range(len(img)):
        zoomFactor = (windowsSize[0] / img[x].shape[0], windowsSize[1] / img[x].shape[1], windowsSize[2] / img[x].shape[2])
        imgResized.append(cpndi.zoom(img[x][:, :, :], zoomFactor, order=interpOrderImg, mode='constant')[:, :, :, cp.newaxis])
        labelResized.append(cpndi.zoom(label[x][:, :, :], zoomFactor, order=interpOrderLabel, mode='constant')[:, :, :, cp.newaxis])
    return imgResized, labelResized

def seed_everything(seed=42):
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    cp.random.seed(seed)
    tf.random.set_random_seed(seed)
    session_conf = tf.compat.v1.ConfigProto(
        intra_op_parallelism_threads=1,
        inter_op_parallelism_threads=1
    )
    sess = tf.compat.v1.Session(graph=tf.compat.v1.get_default_graph(), config=session_conf)
    tf.compat.v1.keras.backend.set_session(sess)

def dice_coef(y_true, y_pred, smooth=1):
    """
    Dice = (2*|X & Y|)/ (|X|+ |Y|)
         =  2*sum(|A*B|)/(sum(A^2)+sum(B^2))
    ref: https://arxiv.org/pdf/1606.04797v1.pdf
    """
    intersection = K.sum(y_true * y_pred)
    return (2. * intersection + smooth) / (K.sum(y_true) + K.sum(y_pred) + smooth)

def dice_loss(y_true, y_pred):
    return 1 - dice_coef(y_true, y_pred)  

def precision(y_true, y_pred, smooth=1):
    # Calculates the precision
    intersection = K.sum(y_true * y_pred)
    return (intersection + smooth) / (K.sum(y_true) + K.sum(y_pred) + smooth - intersection)

def convBlock(x, stage, branch, nb_filter, dropout_rate=None):
    conv_name_base = 'conv' + str(stage) + '_' + str(branch)
    relu_name_base = 'actvi' + str(stage) + '_' + str(branch)
    bn_name_base = 'bn' + str(stage) + '_' + str(branch)
    # 3x3 Convolution
    x = Conv3D(nb_filter, 3, name=conv_name_base, padding="same")(x)
    x = Activation('selu', name=relu_name_base)(x)
    x = BatchNormalization(name=bn_name_base)(x)
    if dropout_rate:
        x = Dropout(dropout_rate)(x)
    return x
    
def create_DenseBlock(inputTensor, branch, stage, n_dense_channels, dilation_rates, dropout_rate):
    concat_feat = inputTensor
    for x in range(n_dense_channels):
        branch = branch + 1
        l = convBlock(concat_feat, branch, stage, n_dense_channels, dropout_rate)
        concat_feat = concatenate([concat_feat, l], axis = -1, name='concat_'+str(stage)+'_'+str(branch))
    return concat_feat

def build_DVN():
    """
    DenseVnet
    """
    main_input = Input(shape=(None, None, None, 1), dtype='float64', name='main_input')
    PreviousSeg_input = Input(shape=(None, None, None, 1), dtype='float64', name='PreviousSeg')
    HintMap_input = Input(shape=(None, None, None, 1), dtype='float64', name='HintMap')
        
    #initialConv = Conv3D(24, 5, strides=(2, 2, 1), padding="same")(main_input)
    initialConv = Conv3D(24, 5, strides=1, padding="same", name='initialConv')(main_input)
    dilation_rates = [5, 10, 10]
    n_dense_channels = [4, 8, 24]
    inLayer = initialConv
    DenseBlock1 = create_DenseBlock(inLayer, 0, 0, n_dense_channels[0], dilation_rates[0], 0.25)
    x = Conv3D(n_dense_channels[0], 3, strides=(2, 2, 1), name='downSampling1', padding="same")(DenseBlock1)
    DenseBlock2 = create_DenseBlock(x, 0, 1, n_dense_channels[1], dilation_rates[1], 0.25)
    y = Conv3D(n_dense_channels[1], 3, strides=(2, 2, 1), name='downSampling2', padding="same")(DenseBlock2)
    DenseBlock3 = create_DenseBlock(y, 0, 2, n_dense_channels[2], dilation_rates[2], 0.25)
    
    skipConv1 = Conv3D(n_dense_channels[0], 3, name='skipConv1', padding="same")(DenseBlock1)
    skipConv2 = Conv3D(n_dense_channels[1], 3, name='skipConv2', padding="same")(DenseBlock2)
    skipConv3 = Conv3D(n_dense_channels[2], 3, name='skipConv3', padding="same")(DenseBlock3)
    
    #upSampling1 = Lambda(.upsampling,arguments={'img_w':skipConv2.get_shape().as_list()[1] * 2,'img_h':skipConv2.get_shape().as_list()[2] * 2})(skipConv2)
    #upSampling2 = Lambda(.upsampling,arguments={'img_w':skipConv2.get_shape().as_list()[1] * 4,'img_h':skipConv2.get_shape().as_list()[2] * 4})(skipConv3)
    upSampling1 = Conv3DTranspose(skipConv2.get_shape().as_list()[-1], 3, strides=(2, 2, 1), padding='same')(skipConv2)
    upSampling2 = Conv3DTranspose(skipConv3.get_shape().as_list()[-1], 3, strides=(4, 4, 1), padding='same')(skipConv3)
   
    connect = concatenate([inLayer, skipConv1, upSampling1, upSampling2], axis = -1, name='concat_main')
    finalConv_0 = Conv3D(1, 3, padding="same", name='finalConv_0')(connect)
    finalActi_0 = Activation('sigmoid', name='final_acti_0')(finalConv_0)
    connect_toInput = concatenate([finalActi_0, PreviousSeg_input, HintMap_input], axis = -1, name='concat_toInput')
    depthWiseConv = Conv3D(3, 3, padding="same", groups=3, name='depthWiseConv')(connect_toInput)
    pointWiseConv = Conv3D(1, 1, padding="same", name='pointWiseConv')(depthWiseConv)
    
    #finalConv_1 = Conv3D(1, 3, padding="same")(connect_toInput)
    finalActi_1 = Activation('sigmoid', name='final_acti_1')(pointWiseConv)
    
    
    model = Model(inputs=[main_input, PreviousSeg_input, HintMap_input], outputs=[finalActi_1])
    return model
    
def file_name(file_dir):   
   label=[] 
   img = []
   path_list = os.listdir(file_dir)
   path_list.sort() 
   for filename in path_list:
       if 'mat' in filename:
           if 'ROI' in filename:
               label.append(os.path.join(filename))
           else:
               img.append(os.path.join(filename))
   return img, label

def meanForAll(x):
    meanList = []
    for img in x:
        meanList.append(cp.mean(img))
    return cp.mean(meanList)

def stdForAll(x):
    stdList = []
    mean = meanForAll(x)
    for img in x:        
        stdList.append((cp.mean(abs(img - mean)**2)))
    return cp.sqrt(cp.sum(stdList) / len(x))

def morphologyHintMap(img, kernel = cp.ones((5,5,5))):
    if len(img.shape) == 3:
        ero = cpndi.binary_erosion(img, structure=kernel).astype(cp.float64)
        dli = cpndi.binary_dilation(img, structure=kernel).astype(cp.float64)
    elif len(img.shape) == 4:
        ero = cpndi.binary_erosion(img[:, :, :, 0], structure=kernel).astype(cp.float64)
        dli = cpndi.binary_dilation(img[:, :, :, 0], structure=kernel).astype(cp.float64)
    elif len(img.shape) == 5:
        ero = cpndi.binary_erosion(img[0, :, :, :, 0], structure=kernel).astype(cp.float64)
        dli = cpndi.binary_dilation(img[0, :, :, :, 0], structure=kernel).astype(cp.float64)
    return dli - ero, dli, ero

def generate_mask(img_height,img_width,img_depth,radius,center_x, center_y, center_z):
    x = cp.array(list(range(img_height))).reshape([img_height,1,1])
    y = cp.array(list(range(img_width))).reshape([1,img_width,1])
    z = cp.array(list(range(img_depth))).reshape([1,1,img_depth])
    # circle mask
    mask = (x-center_x)**2+(y-center_y)**2+(z-center_z)**2<=radius**2  
    return cp.where(mask == True, 1, 0)

def maxKernelRadius(img, label):
    candidate = list(range(3, 41, 2))
    left = 0
    right = len(candidate) - 1
    while(right - left > 1):
        mid = (right + left) // 2
        kernel  = generate_mask(candidate[mid], candidate[mid], candidate[mid], (candidate[mid] - 1) // 2, (candidate[mid] - 1) // 2, (candidate[mid] - 1) // 2, (candidate[mid] - 1) // 2)
        if cp.all(cpndi.binary_erosion(label, structure=kernel).astype(cp.float64) == 0):
            right = mid
        else:
            left = mid
    if cp.all(cpndi.binary_erosion(label, structure=kernel).astype(cp.float64) == 0):
        return candidate[left]
    else:
        return candidate[right]

def cart2pol(x, y):
    rho = np.sqrt(x**2 + y**2)
    phi = np.arctan2(y, x)
    return(rho, phi)
    
def pol2cart(rho, phi):
    x = rho * np.cos(phi)
    y = rho * np.sin(phi)
    return(x, y)

def sortCoord(coords, size):
    pol = []
    cart = []
    for p in coords:
        t = (p[0] - (size - 1) / 2, -1 * p[1] + (size - 1) / 2)
        p_pol = cart2pol(t[0], t[1])
        pol.append(p_pol)
    pol_sorted = sorted(pol , key=lambda k: [k[1], k[0]], reverse=True)
    for p in pol_sorted:
        p_cart = pol2cart(p[0], p[1])
        p_ori_cart = ((int)(np.round(p_cart[0] + (size - 1) / 2)), (int)(np.round(-1 * (p_cart[1] - (size - 1) / 2))))
        cart.append(p_ori_cart)
    return cart

def randomKernel(size = 5):
    k = []
    for z in range(size):
        points = random.randint(4,8)
        polygon = []
        for points in range(points):
            x = random.randint(0,size)
            y = random.randint(0,size)
            polygon.append((x, y))
        polygon_sorted = sortCoord(polygon, size)
        i = Image.new('L', (size, size), 0)
        ImageDraw.Draw(i).polygon(polygon_sorted, outline=1, fill=1)
        mask = np.array(i)
        k.append(mask[:, :, np.newaxis])
    kernel = np.concatenate(k, axis = -1)      
    return kernel

def morphological_augmentation(img, label, kernel):
    resultImg = []
    resultLabel = []
    for x in range(len(img)):
        #kernel = generate_mask(size, size, size, (size - 1) // 2, (size - 1) // 2, (size - 1) // 2, (size - 1) // 2)
        indexList = random.sample(list(range(len(img))), 2)
        indexA = indexList[0]
        indexC = indexList[1]
        hm, d, e = morphologyHintMap(label[x], kernel = cp.asarray(kernel))
        #alpha = random.uniform(0.5, 1.0)
        replacedE = e * label[indexA]
        replacedBackground = (1 - d) * (1 - label[indexC])
        leftRegion = (1 - replacedE - replacedBackground - hm)
        result = (replacedBackground * img[indexC] + (hm + leftRegion) * img[x] + replacedE * img[indexA])
        resultImg.append(cpndi.gaussian_filter(result, 3)[:, :, :, cp.newaxis])
        resultLabel.append(label[x][:, :, :, cp.newaxis])
    return resultImg, resultLabel

def augmentation(img, label, rotation = [-10, 10], rescale = [-10, 10], flip = (1,)):
    flipLayer = RandomFlipLayer(flip_axes = flip)
    rotationLayer = RandomRotationLayer(uniform_angle = True)
    scalingLayer = RandomSpatialScalingLayer()
    im = copy.deepcopy(img)
    la = copy.deepcopy(label)
    return rotationLayer.layer_op(scalingLayer.layer_op(flipLayer.layer_op(im))), rotationLayer.layer_op(scalingLayer.layer_op(flipLayer.layer_op(la), isLabel = True), isLabel = True)

def preprocessing(img, label, sigma, beta, mean, dev, kernel, windowsSize = (40, 40, 40), interpOrderImg = 3, interpOrderLabel = 0):
    imgResized = []
    labelResized = []
    #resize to 40^3    
    im = copy.deepcopy(img)
    la = copy.deepcopy(label)
    for x in range(len(im)):
        im[x] = (im[x] - mean) / dev
    augedImg, augLabel = augmentation(im, la)
    for x in range(len(im)):
        zoomFactor = (windowsSize[0] / augedImg[x].shape[0], windowsSize[1] / augedImg[x].shape[1], windowsSize[2] / augedImg[x].shape[2])
        imgResized.append((cpndi.zoom(augedImg[x][:, :, :, 0], zoomFactor, order=interpOrderImg, mode='constant') - beta * cpndi.gaussian_laplace(cpndi.zoom(augedImg[x][:, :, :, 0], zoomFactor, order=interpOrderImg, mode='constant'), sigma=sigma))[:, :, :, cp.newaxis])
        labelResized.append(cpndi.zoom(augLabel[x][:, :, :, 0], zoomFactor, order=interpOrderLabel, mode='constant')[:, :, :, cp.newaxis])
    #morImg, morLabel = morphological_augmentation(imgResized, labelResized, kernel)
    return imgResized, labelResized

def preprocessing_val(im, la, sigma, beta, mean, dev, windowsSize = (40, 40, 40), interpOrderImg = 3, interpOrderLabel = 0):
    imgResized = []
    labelResized = []    
    img = copy.deepcopy(im)
    label = copy.deepcopy(la)
    for x in range(len(img)):
        img[x] = (img[x] - mean) / dev
    for x in range(len(img)):
        zoomFactor = (windowsSize[0] / img[x].shape[0], windowsSize[1] / img[x].shape[1], windowsSize[2] / img[x].shape[2])
        imgResized.append((cpndi.zoom(img[x], zoomFactor, order=interpOrderImg, mode='constant') - beta * cpndi.gaussian_laplace(cpndi.zoom(img[x], zoomFactor, order=interpOrderImg, mode='constant'), sigma=sigma))[:, :, :, cp.newaxis])
        labelResized.append(cpndi.zoom(label[x], zoomFactor, order=interpOrderLabel, mode='constant')[:, :, :, cp.newaxis])
    return imgResized, labelResized

def morphologyHintLoss(y_true, y_pred):
    hintMap = morphologyHintMap(y_pred, kernel = cp.ones((5,5,5)))
    return K.log(K.sum(hintMap * y_true * y_pred) / K.sum(hintMap * y_pred))

def trainingQueue(imgs, label, queue, sigma, beta, mean, dev, kernel, sample_per_volume = 1, queue_length = 8, batch_size = 16, epoch = 3000000):    
    while(len(imgs) <= batch_size):
        imgs = imgs + imgs
    
    for e in range(epoch):
        temp = list(zip(imgs, label))
        random.shuffle(temp)
        imgs, label = zip(*(temp))
        for x in range(0, len(imgs), batch_size // sample_per_volume):
            tempSampleList = []
            tempLabelList = []
            for t in range(x,  min(x + batch_size // sample_per_volume, len(imgs))):
                for _ in range(sample_per_volume):
                    tempSampleList.append(imgs[t])
                    tempLabelList.append(label[t])
            tempSampleList, tempLabelList = preprocessing(tempSampleList, tempLabelList, sigma, beta, mean, dev, kernel)
            if len(tempSampleList) == batch_size:
                for Sample, l in zip(tempSampleList, tempLabelList):
                    if queue.is_full() is False:
                        queue.push([Sample, l])
                    else:
                        queue.shuffle()
                        yield queue
def main():    
    file = '..//train//'
    val = '..//validate//'
    
    trainImgName, trainLabelName = file_name(file)
    trainImgList, trainlabelList = [], []
    
    valImgName, valLabelName = file_name(val)
    valImgList, vallabelList = [], []
    '''
    Flip = RandomFlipLayer(flip_axes=[1])
    Flip.randomise()
    
    Scaling = RandomSpatialScalingLayer(min_percentage=-10, max_percentage=10)
    Scaling.randomise()
    
    Rotating = RandomRotationLayer()
    Rotating.init_uniform_angle((-10, 10))
    Rotating.randomise()
    '''
    
    for i in range(len(trainImgName)):
        #trainImg = sitk.ReadImage(file + trainImgName[i], sitk.sitkFloat64)
        #trainLabelImg = sitk.ReadImage(file + trainLabelName[i], sitk.sitkFloat64)
        trainImg = scio.loadmat(os.path.join(file, trainImgName[i]))['q_img_c'].astype(np.float64)
        trainLabelImg = scio.loadmat(os.path.join(file, trainLabelName[i]))['ROI_bin_c'].astype(np.float64)
        
        #t = sitk.GetArrayFromImage(trainImg)
        #a = sitk.GetArrayFromImage(trainLabelImg)
        trainImgList.append(cp.asarray(trainImg))
        trainlabelList.append(cp.asarray(trainLabelImg))
    
    allTrainDate = cp.concatenate(trainImgList, axis = -1)
    mean = cp.mean(allTrainDate)
    dev = cp.std(allTrainDate)
    
    size = 5
    #kernel = randomKernel(size = 5)
    kernel = generate_mask(size, size, size, (size - 1) // 2, (size - 1) // 2, (size - 1) // 2, (size - 1) // 2)    
    
    print('mean: %f, dev: %f' %(mean, dev))
    
    for i in range(len(valImgName)):
        #valImg = sitk.ReadImage(val + valImgName[i], sitk.sitkFloat64)
        #valLabelImg = sitk.ReadImage(val + valLabelName[i], sitk.sitkFloat64)
        #t = sitk.GetArrayFromImage(valImg)
        #a   = sitk.GetArrayFromImage(valLabelImg)
        
        valImg = scio.loadmat(os.path.join(val, valImgName[i]))['q_img_c'].astype(np.float64)
        valLabelImg = scio.loadmat(os.path.join(val, valLabelName[i]))['ROI_bin_c'].astype(np.float64)
        
        valImgList.append(cp.asarray(valImg))
        vallabelList.append(cp.asarray(valLabelImg))
    
    for lr in [0.01, 0.001, 0.0001]:
        for batch_size in [8, 16, 32, 64]:
            for sigma in [3, 2, 1]:
                for beta in [1, 2, 3]:
                    main_input = Input(shape=(40, 40, 40, 1), dtype='float64', name='main_input')
                    s = Vnet.Vnet3d(main_input)    
                    s.compile(loss=dice_loss,
                            optimizer=RMSprop(lr=lr),
                            metrics=[dice_coef])
                    #result = train(trainSet, hintMapListTr, 2000, ss, s, valSet = valSet, batch_size=1, sample_interval=5)
                    temp = list(zip(trainImgList, trainlabelList))
                    random.shuffle(temp)
                    trainImgList, trainlabelList = zip(*(temp))
                    iteration = 9000
                    queue_length = 8
                    validation_every_n = 25
                    validation_max_iter = 1
                    #save_iter = 3000
                    valLoss_list = []
                    GL_threshold = 2
                    saved = False
                    workbook = openpyxl.Workbook()
                                    
                    queue_length = max(queue_length, batch_size * 5)
                    queue = Stack(length = queue_length)
                    queueItr = trainingQueue(trainImgList, trainlabelList, queue, sigma, beta, mean, dev, kernel)
                    valProcessed, valLabelProcessed = preprocessing_val(valImgList, vallabelList, sigma, beta, mean, dev)
                    for x in range(1, iteration+1):
                        if queue.size() < batch_size:
                            queue = next(queueItr)
                        data = queue.popMore(num = batch_size)
                        trainData = cp.array([i[0] for i in data])
                        trainLabel = cp.array([i[1] for i in data])
                        pre_loss = s.train_on_batch({'main_input':  cp.asnumpy(trainData)}, cp.asnumpy(trainLabel))
                        print("Stage: %s  iteration: %d [DICE loss: %f, IoU.: %f] " % ('worker', x, pre_loss[0], pre_loss[1]))
                        if x != 0:
                            if x % validation_every_n == 0: 
                                for _ in range(validation_max_iter):
                                    valLoss = []                                    
                                    for index in range(len(valProcessed)):
                                        current_Seg = tf.cast(s.predict({'main_input': cp.asnumpy(valProcessed[index][cp.newaxis, :, :, :, :])}), tf.float64)
                                        valLoss.append(dice_loss(cp.asnumpy((valLabelProcessed[index])), current_Seg[0, :, :, :, :]))
                                    meanValLoss = np.mean(valLoss)                        
                                    print("Stage: %s Validation: %d [DICE loss: %f] " % ('worker', x, meanValLoss))
                                valLoss_list.append(meanValLoss)
                                GL = 100 * ((meanValLoss / np.min(valLoss_list)) - 1)                           
                                if meanValLoss == np.min(valLoss_list):
                                    temp = tf.keras.models.clone_model(s)
                                    temp.set_weights(s.get_weights())         
                                    tempIter = x
                                if GL > GL_threshold and np.min(valLoss_list) < 0.2:
                                    excel = [["iterations", "val_loss"]]
                                    for l in range(len(valLoss_list)):
                                        excel.append([(l + 1) * validation_every_n, valLoss_list[l]])
                                    sheet = workbook.create_sheet()
                                    for i in range(0, len(valLoss_list) + 1):
                                        for j in range(0, len(excel[i])):
                                            sheet.cell(row=i+1, column=j+1, value=excel[i][j])
                                    workbook.save('save//excel//optiming_conventional//iter_' + str(x) + '_lr_'+ str(lr) + '_batchsize_'+ str(batch_size) + '_sigma_'+ str(sigma) + '_beta_'+ str(beta) + '_valLoss_conventional.xlsx')                                    
                                    temp.save('save//LoG//optiming_conventional//Worker_iter_' + str(x) + '_lr_'+ str(lr) + '_batchsize_'+ str(batch_size) + '_sigma_'+ str(sigma) + '_beta_'+ str(beta) + '.h5')
                                    saved = True
                                    del temp
                                    gc.collect()
                                    break
                                    
                            #if x % save_iter == 0: 
                                 #s.save('save//LoG//optiming_conventional//Worker_iter_' + str(x) + '_lr_'+ str(lr) + '_batchsize_'+ str(batch_size) + '_sigma_'+ str(sigma) + '_beta_'+ str(beta) + '.h5')
                                
                        mempool.free_all_blocks()
                        pinned_mempool.free_all_blocks()
                    if saved != True:
                        excel = [["iterations", "val_loss"]]
                        for l in range(len(valLoss_list)):
                            excel.append([(l + 1) * validation_every_n, valLoss_list[l]])
                        sheet = workbook.create_sheet()
                        for i in range(0, len(valLoss_list) + 1):
                            for j in range(0, len(excel[i])):
                                sheet.cell(row=i+1, column=j+1, value=excel[i][j])
                        temp.save('save//LoG//optiming_conventional//Worker_iter_' + str(tempIter) + '_lr_'+ str(lr) + '_batchsize_'+ str(batch_size) + '_sigma_'+ str(sigma) + '_beta_'+ str(beta) + '.h5')
                        workbook.save('save//excel//optiming_conventional//iter_' + str(x) + '_lr_'+ str(lr) + '_batchsize_'+ str(batch_size) + '_sigma_'+ str(sigma) + '_beta_'+ str(beta) + '_valLoss_conventional.xlsx')       
                    
main()