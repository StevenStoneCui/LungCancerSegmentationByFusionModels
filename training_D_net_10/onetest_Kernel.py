# -*- coding: utf-8 -*-
"""
Created on Mon Sep 27 14:21:33 2021

@author: ariken
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Apr 19 18:19:34 2021

@author: ariken
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Apr 19 17:13:31 2021

@author: ariken
"""
import random
import tensorflow as tf
import scipy.io as scio
from PIL import Image
from tensorflow.keras.models import load_model
from tensorflow.keras.optimizers import RMSprop
import tensorflow.keras.backend as K
import os 
import SimpleITK as sitk
import copy
import warnings
import cv2
import openpyxl
import cupyx.scipy.ndimage as cpndi
import cupy as cp
import numpy as np

warnings.filterwarnings('ignore', '.*output shape of zoom.*')
os.environ["CUDA_VISIBLE_DEVICES"] = "0"

config = tf.compat.v1.ConfigProto()  
config.gpu_options.allow_growth=True   
session = tf.compat.v1.Session(config=config)
tf.compat.v1.keras.backend.set_session(session)
#tf.config.run_functions_eagerly(True)
mempool = cp.get_default_memory_pool()
pinned_mempool = cp.get_default_pinned_memory_pool()

def toList(array):
    temp = []
    for x in range(array.shape[-2]):
        temp.append(Image.fromarray(cv2.cvtColor(array[:, :, x, :], cv2.COLOR_BGR2RGB), mode = 'RGB'))
    return temp

def dice_coef(y_true, y_pred, smooth=1):
    """
    Dice = (2*|X & Y|)/ (|X|+ |Y|)
         =  2*sum(|A*B|)/(sum(A^2)+sum(B^2))
    ref: https://arxiv.org/pdf/1606.04797v1.pdf
    """
    intersection = K.sum(y_true * y_pred)
    return (2. * intersection + smooth) / (K.sum(y_true) + K.sum(y_pred) + smooth)

def dice_loss(y_true, y_pred):
    return 1 - dice_coef(y_true, y_pred)  

def Precision(y_true, y_pred):
    return K.sum(y_true * y_pred)/ K.sum(y_pred)
    
def Recall(y_true, y_pred):
    return K.sum(y_true * y_pred)/ K.sum(y_true)
    
def generate_mask(img_height,img_width,img_depth,radius,center_x, center_y, center_z):
    x = cp.array(list(range(img_height))).reshape([img_height,1,1])
    y = cp.array(list(range(img_width))).reshape([1,img_width,1])
    z = cp.array(list(range(img_depth))).reshape([1,1,img_depth])
    # circle mask
    mask = (x-center_x)**2+(y-center_y)**2+(z-center_z)**2<=radius**2  
    return cp.where(mask == True, 1, 0)

def file_name(file_dir):   
   label=[] 
   img = []
   path_list = os.listdir(file_dir)
   path_list.sort() 
   for filename in path_list:
       if 'mat' in filename:
           if 'ROI' in filename:
               label.append(os.path.join(filename))
           else:
               img.append(os.path.join(filename))
   return img, label

def morphological_augmentation(img, label):
    resultImg = []
    resultLabel = []
    for x in range(len(img)):
        size = 5
        indexList = random.sample(list(range(len(img))), 2)
        indexA = indexList[0]
        indexC = indexList[1]
        kernel = generate_mask(size, size, size, (size - 1) // 2, (size - 1) // 2, (size - 1) // 2, (size - 1) // 2)        
        hm, d, e = morphologyHintMap(label[x], kernel = cp.asarray(kernel))
        #alpha = random.uniform(0.5, 1.0)
        replacedE = e * label[indexA]
        replacedBackground = (1 - d) * (1 - label[indexC])
        leftRegion = (1 - replacedE - replacedBackground - hm)
        result = (replacedBackground * img[indexC] + (hm + leftRegion) * img[x] + replacedE * img[indexA])
        
        #resultImg.append(result[:, :, :, cp.newaxis])cpndi.
        resultImg.append(cpndi.gaussian_filter(result, sigma = 3)[:, :, :, cp.newaxis])
        resultLabel.append(label[x][:, :, :, cp.newaxis])
    return resultImg, resultLabel

def preprocessing(img, label, sigma, beta, windowsSize = (40, 40, 40), interpOrderImg = 3, interpOrderLabel = 0):
    imgFinal = []
    labelFinal = []
    #resize to 40^3    
    im = copy.deepcopy(img)
    la = copy.deepcopy(label)
    for x in range(len(im)):
        zoomFactor = (windowsSize[0] / im[x].shape[0], windowsSize[1] / im[x].shape[1], windowsSize[2] / im[x].shape[2])
        imgFinal.append((cpndi.zoom(im[x], zoomFactor, order=interpOrderImg, mode='constant') - beta * cpndi.gaussian_laplace(cpndi.zoom(im[x][:, :, :], zoomFactor, order=interpOrderImg, mode='constant'), sigma=sigma)))
        labelFinal.append(cpndi.zoom(la[x], zoomFactor, order=interpOrderLabel, mode='constant'))
        
    morImg, morLabel = morphological_augmentation(imgFinal, labelFinal)    

    return morImg, morLabel

def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet()
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=value[i][j])
    workbook.save(path)
    
def Edge_ExtractFor3(img, pre, pre_final, ref):
    preEdge_RGB = cp.concatenate((cp.zeros(pre.shape)[:, :, :, cp.newaxis], canny3D(pre)[:, :, :, cp.newaxis], cp.zeros(pre.shape)[:, :, :, cp.newaxis]), axis = -1) #predicted one is green
    preEdge_Final_RGB = cp.concatenate((cp.zeros(pre_final.shape)[:, :, :, cp.newaxis], cp.zeros(pre_final.shape)[:, :, :, cp.newaxis], canny3D(pre_final)[:, :, :, cp.newaxis]), axis = -1) #predicted one is green
    refEdge_RGB = cp.concatenate((canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB + refEdge_RGB + preEdge_Final_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed
    
def Edge_Extract(img, pre, ref):
    preEdge_RGB = cp.concatenate((cp.zeros(pre.shape)[:, :, :, cp.newaxis], canny3D(pre)[:, :, :, cp.newaxis], cp.zeros(pre.shape)[:, :, :, cp.newaxis]), axis = -1) #predicted one is green
    refEdge_RGB = cp.concatenate((canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB + refEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extract_cyan(img, pre, ref):
    preEdge_RGB = np.concatenate((canny3D(pre)[:, :, :, np.newaxis], canny3D(pre)[:, :, :, np.newaxis], np.zeros(pre.shape)[:, :, :, np.newaxis]), axis = -1) #predicted one is green
    refEdge_RGB = np.concatenate((canny3D(ref)[:, :, :, np.newaxis], canny3D(ref)[:, :, :, np.newaxis], canny3D(ref)[:, :, :, np.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB + refEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extract_yellow(img, pre, ref):
    preEdge_RGB = np.concatenate((np.zeros(pre.shape)[:, :, :, np.newaxis], canny3D(pre)[:, :, :, np.newaxis], canny3D(pre)[:, :, :, np.newaxis]), axis = -1) #predicted one is green
    refEdge_RGB = np.concatenate((canny3D(ref)[:, :, :, np.newaxis], canny3D(ref)[:, :, :, np.newaxis], canny3D(ref)[:, :, :, np.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB + refEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extract_pink(img, pre, ref):
    preEdge_RGB = np.concatenate((canny3D(pre)[:, :, :, np.newaxis], canny3D(pre)[:, :, :, np.newaxis] * 204 / 255 , canny3D(pre)[:, :, :, np.newaxis]), axis = -1) #predicted one is green
    refEdge_RGB = np.concatenate((canny3D(ref)[:, :, :, np.newaxis], canny3D(ref)[:, :, :, np.newaxis], canny3D(ref)[:, :, :, np.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB + refEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extract_green(img, pre, ref):
    preEdge_RGB = np.concatenate((np.zeros(pre.shape)[:, :, :, np.newaxis], canny3D(pre)[:, :, :, np.newaxis], np.zeros(pre.shape)[:, :, :, np.newaxis]), axis = -1) #predicted one is green
    refEdge_RGB = np.concatenate((canny3D(ref)[:, :, :, np.newaxis], canny3D(ref)[:, :, :, np.newaxis], canny3D(ref)[:, :, :, np.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB + refEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extrac_final(img, pre, ref):
    preEdge_RGB = cp.concatenate((cp.zeros(pre.shape)[:, :, :, cp.newaxis], cp.zeros(pre.shape)[:, :, :, cp.newaxis], canny3D(pre)[:, :, :, cp.newaxis]), axis = -1)
    refEdge_RGB = cp.concatenate((canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB + refEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extract_onlyRef(img, ref):
    refEdge_RGB = cp.concatenate((canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + refEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extract_Ref(ref):
    refEdge_RGB = cp.concatenate((canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis]), axis = -1) #reference one is white
    refEdge_RGB[refEdge_RGB > 255] = 255 #revise the voxels value that are bigger than 255
    refEdge_RGB[refEdge_RGB < 0] = 0
    return refEdge_RGB

def Edge_Extract_onlyPred(img, pre):
    preEdge_RGB = cp.concatenate((cp.zeros(pre.shape)[:, :, :, cp.newaxis], canny3D(pre)[:, :, :, cp.newaxis], cp.zeros(pre.shape)[:, :, :, cp.newaxis]), axis = -1)
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extract_onlyPredFinal(img, pre_final):
    preEdge_Final_RGB = cp.concatenate((cp.zeros(pre_final.shape)[:, :, :, cp.newaxis], cp.zeros(pre_final.shape)[:, :, :, cp.newaxis], canny3D(pre_final)[:, :, :, cp.newaxis]), axis = -1) #predicted one is green
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_Final_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def readImg(root):
    imgList = sitk.ReadImage(root, sitk.sitkUInt8)
    return tf.transpose(sitk.GetArrayFromImage(imgList), [2, 1, 0]).numpy() #sitk reader will change the channel as [z, y, x], so we change it back
    
def canny3D(img):
    i = copy.deepcopy(img)
    for x in range(i.shape[-1]):
        i[:, :, x]= cv2.Canny(i[:, :, x], 127, 127)
    return i

def Gray2RGB3D(img):
    i = copy.deepcopy(img)
    if len(i.shape) == 4:
        iRGB= cp.concatenate((cp.zeros(i.shape), cp.zeros(i.shape), cp.zeros(i.shape)), axis = -1)    
        for x in range(i.shape[-2]):
            iRGB[:, :, x, :]= cv2.cvtColor(i[:, :, x, :], cv2.COLOR_GRAY2RGB)
    elif len(i.shape) == 3:
        iRGB= np.concatenate((np.zeros(i.shape)[:, :, :, np.newaxis], np.zeros(i.shape)[:, :, :, np.newaxis], np.zeros(i.shape)[:, :, :, np.newaxis]), axis = -1)    
        for x in range(i.shape[-1]):
            iRGB[:, :, x, :]= cv2.cvtColor(i[:, :, x], cv2.COLOR_GRAY2RGB)
    elif len(i.shape) == 2:
        iRGB = cv2.cvtColor(i, cv2.COLOR_GRAY2RGB)            
    return iRGB

def resizing(im, la, sigma, beta, mean, dev, windowsSize = (40, 40, 40), interpOrderImg = 3, interpOrderLabel = 0):
    imgResized = []
    labelResized = []    
    img = copy.deepcopy(im)
    label = copy.deepcopy(la)
    for x in range(len(img)):
        img[x] = (img[x] - mean) / dev
    for x in range(len(img)):
        zoomFactor = (windowsSize[0] / img[x].shape[0], windowsSize[1] / img[x].shape[1], windowsSize[2] / img[x].shape[2])
        imgResized.append((cpndi.zoom(img[x], zoomFactor, order=interpOrderImg, mode='constant') - beta * cpndi.gaussian_laplace(cpndi.zoom(img[x], zoomFactor, order=interpOrderImg, mode='constant'), sigma=sigma))[:, :, :, cp.newaxis])
        labelResized.append(cpndi.zoom(label[x], zoomFactor, order=interpOrderLabel, mode='constant')[:, :, :, cp.newaxis])
    return imgResized, labelResized

def resizing_pro(im, la, sigma, beta, mean, dev, windowsSize = (40, 40, 40), interpOrderImg = 3, interpOrderLabel = 0):
    imgResized = []
    labelResized = []    
    img = copy.deepcopy(im)
    label = copy.deepcopy(la)
    for x in range(len(img)):
        img[x] = (img[x] - mean) / dev
    for x in range(len(img)):
        zoomFactor = (windowsSize[0] / img[x].shape[0], windowsSize[1] / img[x].shape[1], windowsSize[2] / img[x].shape[2])
        imgResized.append(cpndi.gaussian_filter((cpndi.zoom(img[x], zoomFactor, order=interpOrderImg, mode='constant') - beta * cpndi.gaussian_laplace(cpndi.zoom(img[x], zoomFactor, order=interpOrderImg, mode='constant'), sigma=sigma)), 3)[:, :, :, cp.newaxis])
        labelResized.append(cpndi.zoom(label[x], zoomFactor, order=interpOrderLabel, mode='constant')[:, :, :, cp.newaxis])
    return imgResized, labelResized

def computeQualityMeasures(lP,lT):
    quality=[]
    labelPred = lP
    labelTrue = lT
    hausdorffcomputer=sitk.HausdorffDistanceImageFilter()
    hausdorffcomputer.Execute(sitk.GetImageFromArray(labelTrue),sitk.GetImageFromArray(labelPred))
    quality.append(hausdorffcomputer.GetHausdorffDistance()) 
    return quality


val = '..//data_20//test_10//'    
savePNG = ['oneTest//kernel//sphere+random//', 'oneTest//kernel//sphere//','oneTest//kernel//random//','oneTest//kernel//conventional//']
saveVideo = 'test//LoG//Video//'
savePNG_slices = 'PNG//'
model_path_sphereRandom = "save//LoG//optiming_sphere+random_replace//Worker_iter_5850_lr_0.01_batchsize_8_sigma_1_beta_3.h5"
model_path_sphere = "save//LoG//optiming_sphere_replace//Worker_iter_450_lr_0.001_batchsize_16_sigma_1_beta_1.h5"
model_path_Random = "save//LoG//optiming_random_replace//Worker_iter_2700_lr_0.0001_batchsize_32_sigma_1_beta_3.h5"
model_path_conventional = "save//LoG//optiming_conventional//Worker_iter_2975_lr_0.001_batchsize_8_sigma_2_beta_3.h5"
exlPath = "//"
valImgName, valLabelName = file_name(val)
valImgList, vallabelList = [], []
mean= 56.92851563
dev= 41.0767481
windowsSize = (40, 40, 40)


for i in range(len(valImgName)):
    #trainImg = sitk.ReadImage(file + trainImgName[i], sitk.sitkFloat64)
    #trainLabelImg = sitk.ReadImage(file + trainLabelName[i], sitk.sitkFloat64)
    valImg = scio.loadmat(os.path.join(val, valImgName[i]))['q_img_c'].astype(np.float64)
    valLabelImg = scio.loadmat(os.path.join(val, valLabelName[i]))['ROI_bin_c'].astype(np.float64)
    
    #t = sitk.GetArrayFromImage(trainImg)
    #a = sitk.GetArrayFromImage(trainLabelImg)
    valImgList.append(cp.asarray(valImg))
    vallabelList.append(cp.asarray(valLabelImg))
    

workbook = openpyxl.Workbook()
count = 1
edge_extraction = [Edge_Extract_cyan ,Edge_Extract_green, Edge_Extract_yellow, Edge_Extract_pink]
modelIndex = 4
for model_path in [model_path_conventional]:
    dice = []
    precision = []
    recall = []
    excel = [["caseID", "DSC"]]    
    predPre = []
    haus = []
    edged = []
    HM = []
    HMLabel = []
    dliList = []
    eroList = []
    dliListLabel = []
    eroListLabel = []
    reversedDliList = []
    reversedDliLabelList = []
    maxSizeVal = []
    morphoDataList = []
    predFinal = []
    dicePre = []
    diceFinal = []
        
    sigma = float(model_path.replace(".h5","").split("sigma_")[1].split("_")[0])  
    beta = float(model_path.replace(".h5","").split("beta_")[1].split("_")[0])
    
    
    if modelIndex == 4:
        valProcessed, valLabelProcessed = resizing(valImgList, vallabelList, sigma, beta, mean, dev)
    else:
        valProcessed, valLabelProcessed = resizing_pro(valImgList, vallabelList, sigma, beta, mean, dev)
    
    model = load_model(model_path, compile = False)
    model.compile(loss=dice_loss,
            optimizer=RMSprop(lr=0.001),
            metrics=[dice_coef])    

    
    index = 0
    p = model.predict({'main_input': cp.asnumpy(cp.asarray(valProcessed))})
    p[p > 0.5] = 1
    p[p <= 0.5] = 0
    for x in range(len(valProcessed)):  
        zoomFactor = (windowsSize[0] / valImgList[x].shape[0], windowsSize[1] / valImgList[x].shape[1], windowsSize[2] / valImgList[x].shape[2])        
        v = cpndi.zoom(valImgList[x], zoomFactor, order=3, mode='constant')[:, :, :]
        v[v>255] = 255
        v[v<0] = 0
        t = edge_extraction[count - 1](cp.asnumpy(v.astype(np.float32)), p[x, :, :, :, 0].astype(np.uint8)*255, cp.asnumpy(valLabelProcessed[x][:, :, :, 0]).astype(np.uint8)*255)
        cv2.imwrite(savePNG[count - 1] + valImgName[x].split('.nii')[0] + '_edged_10.png', t[:, :, 10])
        cv2.imwrite(savePNG[count - 1] + valImgName[x].split('.nii')[0] + '_10.png', cp.asnumpy(v[:, :, 10]))
        cv2.imwrite(savePNG[count - 1] + valImgName[x].split('.nii')[0] + '_edged_20.png', t[:, :, 20])
        cv2.imwrite(savePNG[count - 1] + valImgName[x].split('.nii')[0] + '_20.png', cp.asnumpy(v[:, :, 20]))
        cv2.imwrite(savePNG[count - 1] + valImgName[x].split('.nii')[0] + '_edged_30.png', t[:, :, 30])
        cv2.imwrite(savePNG[count - 1] + valImgName[x].split('.nii')[0] + '_30.png', cp.asnumpy(v[:, :, 30]))
        dice.append(float(dice_coef(p[x, :, :, :, 0], cp.asnumpy(cp.asarray(valLabelProcessed[x][:, :, :, 0].astype(np.float32))))))
        precision.append(float(Precision(cp.asnumpy(cp.asarray(valLabelProcessed[x][:, :, :, 0].astype(np.float32))), p[x, :, :, :, 0])))
        recall.append(float(Recall(cp.asnumpy(cp.asarray(valLabelProcessed[x][:, :, :, 0].astype(np.float32))), p[x, :, :, :, 0])))
        excel.append([valImgName[x], dice[-1], precision[-1], recall[-1]])
    print(str(count), model_path.split('//')[-1] + '\t', float(np.mean(dice)))
    excel.append(['', '=AVERAGE(B2:B165)', '=AVERAGE(C2:C165)', '=AVERAGE(D2:D165)'])
    count = count + 1
    index = len(excel)
    sheet = workbook.create_sheet(title=model_path.split('//')[-2], index = 0)
    for i in range(0, index):
        for j in range(0, len(excel[i])):
            sheet.cell(row=i+1, column=j+1, value=excel[i][j])
    modelIndex = modelIndex + 1
workbook.save('onetest_kernel.xlsx')

'''
    excel = [["caseID", "DSC of without morphology"]]
    
    for x in range(len(dicePre)):
        excel.append([valImgName[x], dicePre[x]])
            
    for weights in range(10):
        diceFinal = []
        index = 0
        excel[0].append("DSC of with morphology, w = %d" %(weights))
        for x in morphoDataList:
            temp = cp.zeros((40, 40, 40))
            factor = 0
            for img in x:
                p = model1.predict({'main_input': img[cp.newaxis, :, :, :, :]})
                r = p[0, :, :, :, 0]
                temp = temp + r
                factor = factor + 1
            result = (cp.round((temp + predPre[index] * weights) / (factor + weights))).astype(cp.float32)
            #result = cp.where((temp / factor) >= 2/3, 1, 0).astype(cp.float32)
            diceFinal.append(dice_coef(valLabelProcessed[index][:, :, :, 0], result, smooth = 0.000001).numpy())
            predFinal.append(result[cp.newaxis, :, :, :, cp.newaxis])
            index = index + 1
            
        for x in range(1, len(excel)):
            excel[x].append(diceFinal[x - 1])
        
    write_excel_xlsx("F:\\Cui\\RL+DVN\\TeacherTraining\\FinalResult.xlsx", 'result', excel)
'''
'''
    index = 0
    for x in morphoDataList:
        temp = cp.zeros((40, 40, 40))
        factor = 0
        for img in x:
            p = model1.predict({'main_input': img[cp.newaxis, :, :, :, :]})
            r = p[0, :, :, :, 0]
            temp = temp + r
            factor = factor + 1
        result = (cp.round((temp + predPre[index] * factor) / (factor + factor))).astype(cp.float32)
        diceFinal.append(dice_coef(valLabelProcessed[index][:, :, :, 0], result, smooth = 0.000001).numpy())
        predFinal.append(result[cp.newaxis, :, : , :, cp.newaxis])
        #predFinal.append(cp.where(temp >= factor*2//3, 1, 0))
        index = index + 1
        
    excel = [["caseID", "DSC of without morphology", "DSC of with morphology"]]
    
    for x in range(len(dicePre)):
        excel.append([valImgName[x], dicePre[x], diceFinal[x]])
        
    write_excel_xlsx("F:\\Cui\\RL+DVN\\TeacherTraining\\FinalResult.xlsx", 'result', excel)
'''
    
        
'''
pred = predFinal
for im in range(len(pred)):
    kernel = generate_mask(7, 7, 7, 3, 3, 3, 3)
    #kernel = generate_mask(5, 5, 5, 2, 2, 2, 2)
    #kernel = generate_mask(3, 3, 3, 1, 1, 1, 1)
    hm, d, e = morphologyHintMap(cp.where(pred[im]>0.5, 1, 0), kernel = kernel)
    HM.append(hm)
    dliList.append(d)
    eroList.append(e)
    cv2.imwrite(savePNG + valImgName[im] + '_hintmap.png', HM[im][:, :, HM[im].shape[-1] // 2] * 255)
    cv2.imwrite(savePNG + valImgName[im] + '_pred.png', pred[im][0, :, :, HM[im].shape[-1] // 2, 0] * 255)
    hm, d, e = morphologyHintMap(valLabelProcessed[im][:, :, :, 0], kernel = kernel)
    HMLabel.append(hm)
    dliListLabel.append(d)
    eroListLabel.append(e)
    cv2.imwrite(savePNG + valImgName[im] + '_hintmap_label.png', HMLabel[im][:, :, HM[im].shape[-1] // 2] * 255)
    cv2.imwrite(savePNG + valImgName[im] + '_label.png', valLabelProcessed[im][:, :, HM[im].shape[-1] // 2, 0] * 255)
    cv2.imwrite(savePNG + valImgName[im] + '_dli_label.png', dliListLabel[im][:, :, 20] * 255)
    cv2.imwrite(savePNG + valImgName[im] + '_ero_label.png', eroListLabel[im][:, :, 20] * 255)
    
meanTumorB, meanBack = toGetBand.FillMean()
meanBackAvg = meanBack

for x in range(len(dliList)):
    reversedDliList.append(1 - dliList[x])
    reversedDliLabelList.append(1 - dliListLabel[x])
    meanBack = min(cp.min(valProcessed[x]), meanBack)*56.15427+71.85165
    m = cp.mean([valProcessed[x][HM[x].shape[0] // 2, HM[x].shape[1] // 2, HM[2].shape[-1] // 2, 0], \
                valProcessed[x][-1 + HM[x].shape[0] // 2, HM[x].shape[1] // 2, HM[2].shape[-1] // 2, 0], \
                valProcessed[x][-1 + HM[x].shape[0] // 2, -1 + HM[x].shape[1] // 2, HM[2].shape[-1] // 2, 0], \
                valProcessed[x][-1 + HM[x].shape[0] // 2, -1 + HM[x].shape[1] // 2, -1 + HM[2].shape[-1] // 2, 0], \
                valProcessed[x][-1 + HM[x].shape[0] // 2, HM[x].shape[1] // 2, -1 + HM[2].shape[-1] // 2, 0], \
                valProcessed[x][HM[x].shape[0] // 2, -1 + HM[x].shape[1] // 2, HM[2].shape[-1] // 2, 0], \
                valProcessed[x][HM[x].shape[0] // 2, -1 + HM[x].shape[1] // 2, -1 + HM[2].shape[-1] // 2, 0], \
                valProcessed[x][HM[x].shape[0] // 2, HM[x].shape[1] // 2, -1 + HM[2].shape[-1] // 2, 0]])
    meanTumor = max(m, meanTumorB)*56.15427+71.85165
    revised = HM[x] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0] + reversedDliList[x] * meanBack + eroList[x] * meanTumor * 0.5 + eroList[x] * 0.5 * (valProcessed[x] * 56.15427+71.85165)[:, :, :, 0]
    revisedlabel = HMLabel[x] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0] + reversedDliLabelList[x] * meanBack + eroListLabel[x] * meanTumor * 0.5 + eroListLabel[x]  * 0.5 * (valProcessed[x] * 56.15427+71.85165)[:, :, :, 0]
    
    revised[revised < 0] = 0
    revisedlabel[revisedlabel < 0] = 0
    revised[revised > 255] = 255
    revisedlabel[revisedlabel > 255] = 255

    notRevised = HM[x][:, :, :] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0]
    notRevisedLabel = HMLabel[x][:, :, :] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0]
    cv2.imwrite(savePNG + valImgName[x] + '_revised.png', revised[:, :, revised.shape[2] // 2])
    cv2.imwrite(savePNG + valImgName[x] + '_revised_label.png', revisedlabel[:, :, revisedlabel.shape[2] // 2])
    t = Edge_Extract((revised).astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_revised&ref&pred.png', t[:, :, revised.shape[2] // 2])
    t = Edge_Extract((revisedlabel).astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_revised&ref&pred_label.png', t[:, :, revised.shape[2] // 2])
    t = Edge_Extract_onlyRef((revised).astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_revised&ref.png', t[:, :, t.shape[-2] // 2, :])
    t = Edge_Extract_onlyPred((revised).astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_revised&pred.png', t[:, :, t.shape[-2] // 2, :])
    nt = Edge_Extract_onlyRef((notRevised).astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_notRevised&ref.png', nt[:, :, nt.shape[-2] // 2, :])
    
    t = Edge_Extrac_final((revised).astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_preFinald&ref&revised.png', t[:, :, t.shape[-2] // 2, :])  
    t = Edge_Extract_onlyRef((revisedlabel).astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    nt = Edge_Extract_onlyRef((notRevisedLabel).astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_revised&ref_label.png', t[:, :, t.shape[-2] // 2, :])
    cv2.imwrite(savePNG + valImgName[x] + '_notRevised&ref_label.png', nt[:, :, nt.shape[-2] // 2, :])
    #Image.fromarray((HM[x][:, :, HM[x].shape[-1] // 2] * (valProcessed[x]*56.15427+71.85165)[:, :, HM[x].shape[-1] // 2, 0] + reversedDliList[x][:, :, HM[x].shape[-1] // 2] * meanBack + eroList [x][:, :, HM[x].shape[-1] // 2] * meanTumor).astype(cp.uint8)).save(savePNG + valImgName[x] + '_revised.png')
    #imageio.imsave(savePNG + valImgName[x] + '_revised.png', (HM[x][:, :, HM[x].shape[-1] // 2] * (valProcessed[x]*56.15427+71.85165)[:, :, HM[x].shape[-1] // 2, 0] + reversedDliList[x][:, :, HM[x].shape[-1] // 2] * meanBack + eroList [x][:, :, HM[x].shape[-1] // 2] * meanTumor).astype(cp.uint8))

for x in range(len(pred)):
    dice.append(dice_coef(valLabelProcessed[x], pred[x], smooth = 0.000001).numpy())
    haus.append(computeQualityMeasures(valLabelProcessed[x][:, :, :, 0].astype(cp.uint8), pred[x][0, :, :, :, 0].astype(cp.uint8)))
    img8bit = valProcessed[x]*56.15427+71.85165
    img8bit[img8bit < 0] = 0
    img8bit[img8bit > 255] = 255
    cv2.imwrite(savePNG + valImgName[x] + '_origin.png', img8bit[:, :, HM[x].shape[-1] // 2, 0])
    cv2.imwrite(savePNG + valImgName[x] + '_masked.png', HM[x][:, :, HM[x].shape[-1] // 2] * img8bit[:, :, HM[x].shape[-1] // 2, 0])
    cv2.imwrite(savePNG + valImgName[x] + '_masked_label.png', HMLabel[x][:, :, HMLabel[x].shape[-1] // 2] * img8bit[:, :, HMLabel[x].shape[-1] // 2, 0])
    t = Edge_ExtractFor3((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    edged.append(t.astype(cp.uint8))
    edged3D = toList(t.astype(cp.uint8))
    edged3D[0].save(saveGIF + valImgName[x].split(".nii")[0] + '.gif',
           save_all=True, append_images=edged3D[1:], loop = 0, duration = 200)
    hintmap = Gray2RGB3D(HM[x])
    fourcc = cv2.VideoWriter_fourcc('m', 'p', '4', 'v')
    out = cv2.VideoWriter(saveVideo + valImgName[x].split(".nii")[0] + '.mp4',fourcc, 5.0, (40,40))
    for y in range(t.shape[-2]):
        out.write(t.astype(cp.uint8)[:, :, y, :])
    out.release()
    cv2.destroyAllWindows()
    t = Edge_Extract((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_masked&edge.png', hintmap[:, :, hintmap.shape[-2] // 2, :] * t[:, :, t.shape[-2] // 2, :])
    cv2.imwrite(savePNG + valImgName[x] + '_pred&ref.png', t[:, :, t.shape[-2] // 2, :])    
    cv2.imwrite(savePNG + valImgName[x] + '_dli.png', dliList[x][:, :, t.shape[-2] // 2] * 255)
    cv2.imwrite(savePNG + valImgName[x] + '_ero.png', eroList[x][:, :, t.shape[-2] // 2] * 255)
    r = Edge_Extract_onlyRef((img8bit).astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_ref.png', r[:, :, t.shape[-2] // 2, :])
    t = Edge_Extract_onlyPred((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_pred.png', t[:, :, t.shape[-2] // 2, :])
    t = Edge_ExtractFor3((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_pred&predFinal.png', t[:, :, t.shape[-2] // 2, :])
    t = Edge_Extract_onlyPredFinal((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_predFinal.png', t[:, :, t.shape[-2] // 2, :])
    t = Edge_Extract_Ref(255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_refonly.png', t[:, :, t.shape[-2] // 2, :])

for x in range(len(pred)):   
     for slices in range(-2, 6, 1):
         img8bit = valProcessed[x]*56.15427+71.85165
         notRevised = HM[x][:, :, :] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0]
         revised = HM[x] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0] + reversedDliList[x] * meanBack + eroList[x] * meanTumor * 0.5 + eroList[x] * 0.5 * (valProcessed[x] * 56.15427+71.85165)[:, :, :, 0]
         img8bit[img8bit < 0] = 0
         img8bit[img8bit > 255] = 255
         revised[revised < 0] = 0
         revised[revised > 255] = 255
         notRevised[notRevised < 0] = 0
         notRevised[notRevised > 255] = 255
         cv2.imwrite(savePNG_slices + valImgName[x] + '_slices_' + str(slices) + '_origin.png', img8bit[:, :, slices + HM[x].shape[-1] // 2, 0])
         t = Edge_Extract_onlyPred((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8))
         t[t < 0] = 0
         t[t > 255] = 255
         cv2.imwrite(savePNG_slices + valImgName[x] + '_slices_' + str(slices) + '_pred.png', t[:, :, slices + t.shape[-2] // 2, :])
         cv2.imwrite(savePNG_slices + valImgName[x] + '_slices_' + str(slices) + '_revised.png', revised[:, :, slices + revised.shape[2] // 2])
         cv2.imwrite(savePNG_slices + valImgName[x] + '_slices_' + str(slices) + '_notRevised.png', notRevised[:, :, slices + revised.shape[2] // 2])
         t = Edge_ExtractFor3((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
         t[t < 0] = 0
         t[t > 255] = 255
         cv2.imwrite(savePNG_slices + valImgName[x] + '_slices_' + str(slices) + '_pred&predFinal&ref.png', t[:, :, slices + t.shape[-2] // 2, :])
'''