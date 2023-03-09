# -*- coding: utf-8 -*-
"""
Created on Mon Apr 19 18:19:34 2021

@author: ariken
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Apr 19 17:13:31 2021

@author: ariken
"""

import tensorflow as tf
import scipy.io as scio
from PIL import Image
from tensorflow.keras.models import load_model
from tensorflow.keras.optimizers import RMSprop
import tensorflow.keras.backend as K
import os 
import SimpleITK as sitk
import copy
import warnings
import cv2
import openpyxl
import cupyx.scipy.ndimage as cpndi
import cupy as cp
import numpy as np

warnings.filterwarnings('ignore', '.*output shape of zoom.*')
os.environ["CUDA_VISIBLE_DEVICES"] = "0"

config = tf.compat.v1.ConfigProto()  
config.gpu_options.allow_growth=True   
session = tf.compat.v1.Session(config=config)
tf.compat.v1.keras.backend.set_session(session)
#tf.config.run_functions_eagerly(True)
mempool = cp.get_default_memory_pool()
pinned_mempool = cp.get_default_pinned_memory_pool()

def toList(array):
    temp = []
    for x in range(array.shape[-2]):
        temp.append(Image.fromarray(cv2.cvtColor(array[:, :, x, :], cv2.COLOR_BGR2RGB), mode = 'RGB'))
    return temp

def dice_coef(y_true, y_pred, smooth=1):
    """
    Dice = (2*|X & Y|)/ (|X|+ |Y|)
         =  2*sum(|A*B|)/(sum(A^2)+sum(B^2))
    ref: https://arxiv.org/pdf/1606.04797v1.pdf
    """
    intersection = K.sum(y_true * y_pred)
    return (2. * intersection + smooth) / (K.sum(y_true) + K.sum(y_pred) + smooth)

def dice_loss(y_true, y_pred):
    return 1 - dice_coef(y_true, y_pred)  
    
def file_name(file_dir):   
   label=[] 
   img = []
   path_list = os.listdir(file_dir)
   path_list.sort() 
   for filename in path_list:
       if 'mat' in filename:
           if 'ROI' in filename:
               label.append(os.path.join(filename))
           else:
               img.append(os.path.join(filename))
   return img, label

def preprocessing(train, label, downSamplingFactor = 8):
    resultTrain = []
    resultLabel = []
    #mean = meanForAll(train)
    #std = stdForAll(train) 
    for index in range(len(train)):
        #train[index] = (train[index] - cp.mean(train[index]))/cp.std(train[index])
        #train[index] = (train[index] - mean)/std
        mod = train[index].shape[2] % downSamplingFactor
        if mod != 0:
            zeroPre = (train[index].shape[0], train[index].shape[1], int((downSamplingFactor - mod)/2))
            zeroPost = (train[index].shape[0], train[index].shape[1], downSamplingFactor - mod - int((downSamplingFactor - mod)/2))
            resultTrain.append((cp.concatenate((cp.zeros(zeroPre), train[index], cp.zeros(zeroPost)), axis=2))[cp.newaxis, :, :, :, cp.newaxis])
            #resultLabel.append((cp.concatenate((cp.zeros(zeroPre), label[index], cp.zeros(zeroPost)), axis=2))[cp.newaxis, :, :, :, cp.newaxis])
            resultLabel.append(label[index][cp.newaxis, :, :, :, cp.newaxis])
        else:
            resultTrain.append(train[index][cp.newaxis, :, :, :, cp.newaxis])
            resultLabel.append(label[index][cp.newaxis, :, :, :, cp.newaxis])
    return resultTrain, resultLabel

def postprocessing(train, pred, downSamplingFactor = 8):
    resultTrain = []
    for index in range(len(train)):
        #train[index] = (train[index] - cp.mean(train[index]))/cp.std(train[index])
        #train[index] = (train[index] - mean)/std
        mod = train[index].shape[2] % downSamplingFactor
        if mod != 0:
            zeroPre = int((downSamplingFactor - mod)/2)
            resultTrain.append(pred[index][:, :, :, zeroPre+1:train[index].shape[2]+zeroPre+1,:])
            #resultLabel.append((cp.concatenate((cp.zeros(zeroPre), label[index], cp.zeros(zeroPost)), axis=2))[cp.newaxis, :, :, :, cp.newaxis])
        else:
            resultTrain.append(pred[index])
    return resultTrain

def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet()
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=value[i][j])
    workbook.save(path)
    
def Edge_ExtractFor3(img, pre, pre_final, ref):
    preEdge_RGB = cp.concatenate((cp.zeros(pre.shape)[:, :, :, cp.newaxis], canny3D(pre)[:, :, :, cp.newaxis], cp.zeros(pre.shape)[:, :, :, cp.newaxis]), axis = -1) #predicted one is green
    preEdge_Final_RGB = cp.concatenate((cp.zeros(pre_final.shape)[:, :, :, cp.newaxis], cp.zeros(pre_final.shape)[:, :, :, cp.newaxis], canny3D(pre_final)[:, :, :, cp.newaxis]), axis = -1) #predicted one is green
    refEdge_RGB = cp.concatenate((canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB + refEdge_RGB + preEdge_Final_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed
    
def Edge_Extract(img, pre, ref):
    preEdge_RGB = cp.concatenate((cp.zeros(pre.shape)[:, :, :, cp.newaxis], canny3D(pre)[:, :, :, cp.newaxis], cp.zeros(pre.shape)[:, :, :, cp.newaxis]), axis = -1) #predicted one is green
    refEdge_RGB = cp.concatenate((canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB + refEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extrac_final(img, pre, ref):
    preEdge_RGB = cp.concatenate((cp.zeros(pre.shape)[:, :, :, cp.newaxis], cp.zeros(pre.shape)[:, :, :, cp.newaxis], canny3D(pre)[:, :, :, cp.newaxis]), axis = -1)
    refEdge_RGB = cp.concatenate((canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB + refEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extract_onlyRef(img, ref):
    refEdge_RGB = cp.concatenate((canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis]), axis = -1) #reference one is white
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + refEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extract_Ref(ref):
    refEdge_RGB = cp.concatenate((canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis], canny3D(ref)[:, :, :, cp.newaxis]), axis = -1) #reference one is white
    refEdge_RGB[refEdge_RGB > 255] = 255 #revise the voxels value that are bigger than 255
    refEdge_RGB[refEdge_RGB < 0] = 0
    return refEdge_RGB

def Edge_Extract_onlyPred(img, pre):
    preEdge_RGB = cp.concatenate((cp.zeros(pre.shape)[:, :, :, cp.newaxis], canny3D(pre)[:, :, :, cp.newaxis], cp.zeros(pre.shape)[:, :, :, cp.newaxis]), axis = -1)
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def Edge_Extract_onlyPredFinal(img, pre_final):
    preEdge_Final_RGB = cp.concatenate((cp.zeros(pre_final.shape)[:, :, :, cp.newaxis], cp.zeros(pre_final.shape)[:, :, :, cp.newaxis], canny3D(pre_final)[:, :, :, cp.newaxis]), axis = -1) #predicted one is green
    img_RGB = Gray2RGB3D(img)
    img_processed = img_RGB + preEdge_Final_RGB
    img_processed[img_processed > 255] = 255 #revise the voxels value that are bigger than 255
    img_processed[img_processed < 0] = 0
    return img_processed

def readImg(root):
    imgList = sitk.ReadImage(root, sitk.sitkUInt8)
    return tf.transpose(sitk.GetArrayFromImage(imgList), [2, 1, 0]).numpy() #sitk reader will change the channel as [z, y, x], so we change it back
    
def canny3D(img):
    i = copy.deepcopy(img)
    for x in range(i.shape[-1]):
        i[:, :, x]= cv2.Canny(i[:, :, x], 127, 127)
    return i

def Gray2RGB3D(img):
    i = copy.deepcopy(img)
    if len(i.shape) == 4:
        iRGB= cp.concatenate((cp.zeros(i.shape), cp.zeros(i.shape), cp.zeros(i.shape)), axis = -1)    
        for x in range(i.shape[-2]):
            iRGB[:, :, x, :]= cv2.cvtColor(i[:, :, x, :], cv2.COLOR_GRAY2RGB)
    elif len(i.shape) == 3:
        iRGB= cp.concatenate((cp.zeros(i.shape)[:, :, :, cp.newaxis], cp.zeros(i.shape)[:, :, :, cp.newaxis], cp.zeros(i.shape)[:, :, :, cp.newaxis]), axis = -1)    
        for x in range(i.shape[-1]):
            iRGB[:, :, x, :]= cv2.cvtColor(i[:, :, x], cv2.COLOR_GRAY2RGB)
    elif len(i.shape) == 2:
        iRGB = cv2.cvtColor(i, cv2.COLOR_GRAY2RGB)            
    return iRGB

def preprocessing_val(im, la, sigma, beta, mean, dev, windowsSize = (40, 40, 40), interpOrderImg = 3, interpOrderLabel = 0):
    imgResized = []
    labelResized = []    
    img = copy.deepcopy(im)
    label = copy.deepcopy(la)
    for x in range(len(img)):
        img[x] = (img[x] - mean) / dev
    for x in range(len(img)):
        zoomFactor = (windowsSize[0] / img[x].shape[0], windowsSize[1] / img[x].shape[1], windowsSize[2] / img[x].shape[2])
        imgResized.append((cpndi.zoom(img[x], zoomFactor, order=interpOrderImg, mode='constant') - beta * cpndi.gaussian_laplace(cpndi.zoom(img[x], zoomFactor, order=interpOrderImg, mode='constant'), sigma=sigma))[:, :, :, cp.newaxis])
        labelResized.append(cpndi.zoom(label[x], zoomFactor, order=interpOrderLabel, mode='constant')[:, :, :, cp.newaxis])
    return imgResized, labelResized

def preprocessing_pro(im, la, sigma, beta, mean, dev, windowsSize = (40, 40, 40), interpOrderImg = 3, interpOrderLabel = 0):
    imgResized = []
    labelResized = []    
    img = copy.deepcopy(im)
    label = copy.deepcopy(la)
    for x in range(len(img)):
        img[x] = (img[x] - mean) / dev
    for x in range(len(img)):
        zoomFactor = (windowsSize[0] / img[x].shape[0], windowsSize[1] / img[x].shape[1], windowsSize[2] / img[x].shape[2])
        imgResized.append(cpndi.gaussian_filter((cpndi.zoom(img[x], zoomFactor, order=interpOrderImg, mode='constant') - beta * cpndi.gaussian_laplace(cpndi.zoom(img[x], zoomFactor, order=interpOrderImg, mode='constant'), sigma=sigma)), 3)[:, :, :, cp.newaxis])
        labelResized.append(cpndi.zoom(label[x], zoomFactor, order=interpOrderLabel, mode='constant')[:, :, :, cp.newaxis])
    return imgResized, labelResized

def computeQualityMeasures(lP,lT):
    quality=[]
    labelPred = lP
    labelTrue = lT
    hausdorffcomputer=sitk.HausdorffDistanceImageFilter()
    hausdorffcomputer.Execute(sitk.GetImageFromArray(labelTrue),sitk.GetImageFromArray(labelPred))
    quality.append(hausdorffcomputer.GetHausdorffDistance()) 
    return quality

val = '..//data_20//validate_10//'    
saveGIF = 'test//LoG//GIF//'
savePNG = 'test//LoG//PNG//'
saveVideo = 'test//LoG//Video//'
savePNG_slices = 'PNG//'
model_path_sphereRandom = "save//LoG//optiming_sphere+random_replace//"
model_path_sphere = "save//LoG//optiming_sphere_replace//"
model_path_Random = "save//LoG//optiming_random_replace//"
model_path_Conventional = "save//LoG//optiming_conventional//"
exlPath = "//"
valImgName, valLabelName = file_name(val)
valImgList, vallabelList = [], []

file = '..//data_20//train_10//'

trainImgName, trainLabelName = file_name(file)
trainImgList, trainlabelList = [], []

for i in range(len(trainImgName)):
    #trainImg = sitk.ReadImage(file + trainImgName[i], sitk.sitkFloat64)
    #trainLabelImg = sitk.ReadImage(file + trainLabelName[i], sitk.sitkFloat64)
    trainImg = scio.loadmat(os.path.join(file, trainImgName[i]))['q_img_c'].astype(np.float64)
    trainLabelImg = scio.loadmat(os.path.join(file, trainLabelName[i]))['ROI_bin_c'].astype(np.float64)
    
    #t = sitk.GetArrayFromImage(trainImg)
    #a = sitk.GetArrayFromImage(trainLabelImg)
    trainImgList.append(cp.asarray(trainImg))
    trainlabelList.append(cp.asarray(trainLabelImg))
    
allTrainDate = cp.concatenate(trainImgList, axis = -1)
mean = cp.mean(allTrainDate)
dev = cp.std(allTrainDate)

for i in range(len(valImgName)):
    #trainImg = sitk.ReadImage(file + trainImgName[i], sitk.sitkFloat64)
    #trainLabelImg = sitk.ReadImage(file + trainLabelName[i], sitk.sitkFloat64)
    valImg = scio.loadmat(os.path.join(val, valImgName[i]))['q_img_c'].astype(np.float64)
    valLabelImg = scio.loadmat(os.path.join(val, valLabelName[i]))['ROI_bin_c'].astype(np.float64)
    
    #t = sitk.GetArrayFromImage(trainImg)
    #a = sitk.GetArrayFromImage(trainLabelImg)
    valImgList.append(cp.asarray(valImg))
    vallabelList.append(cp.asarray(valLabelImg))
    

workbook = openpyxl.Workbook()
modelIndex = 3
#for model_path in [model_path_sphereRandom, model_path_sphere, model_path_Random, model_path_Conventional]: 
for model_path in [model_path_Conventional]:   
    excel = [["caseID", "DSC"]]
    model_List = os.listdir(model_path)
    model_List.sort()
    model_List.reverse()
    count = 1
    for md in range(len(model_List)): 
        dice = []
            
        sigma = int(model_List[md].replace(".h5","").split("sigma_")[1].split("_")[0])  
        beta = int(model_List[md].replace(".h5","").split("beta_")[1].split("_")[0])
        
        if modelIndex == 3:
            valProcessed, valLabelProcessed = preprocessing_val(valImgList, vallabelList, sigma, beta, mean, dev)
        else:
            valProcessed, valLabelProcessed = preprocessing_pro(valImgList, vallabelList, sigma, beta, mean, dev)
        
        model = load_model(model_path + model_List[md])
        for x in range(len(valProcessed)):  
            p = tf.cast(model.predict({'main_input': cp.asnumpy((valProcessed[x])[cp.newaxis, :, :, :, :])}), tf.float64)      
            p_r = np.where(p>=0.5, 1.0, 0.0)
            dice.append(dice_coef(cp.asnumpy(valLabelProcessed[x]), p_r[0, :, :, :, :]))
        excel.append([model_List[md], np.mean(dice)])        
        print(str(count) + "/" + str(len(model_List)), model_List[md] + '\t', float(np.mean(dice)))
        count = count + 1
    index = len(excel)
    sheet = workbook.create_sheet(title=model_path.split('//')[-2], index = 0)
    modelIndex = modelIndex + 1
    for i in range(0, index):
        for j in range(0, len(excel[i])):
            sheet.cell(row=i+1, column=j+1, value=excel[i][j])
workbook.save('kernelResult_val_conventional.xlsx')

val = '..//data_20//test_10//'
saveGIF = 'test//LoG//GIF//'
savePNG = 'test//LoG//PNG//'
saveVideo = 'test//LoG//Video//'
savePNG_slices = 'PNG//'
model_path_sphereRandom = "save//LoG//optiming_sphere+random_replace//"
model_path_sphere = "save//LoG//optiming_sphere_replace//"
model_path_Random = "save//LoG//optiming_random_replace//"
model_path_Conventional = "save//LoG//optiming_conventional//"
exlPath = "//"
valImgName, valLabelName = file_name(val)
valImgList, vallabelList = [], []


for i in range(len(valImgName)):
    #trainImg = sitk.ReadImage(file + trainImgName[i], sitk.sitkFloat64)
    #trainLabelImg = sitk.ReadImage(file + trainLabelName[i], sitk.sitkFloat64)
    valImg = scio.loadmat(os.path.join(val, valImgName[i]))['q_img_c'].astype(np.float64)
    valLabelImg = scio.loadmat(os.path.join(val, valLabelName[i]))['ROI_bin_c'].astype(np.float64)
    
    #t = sitk.GetArrayFromImage(trainImg)
    #a = sitk.GetArrayFromImage(trainLabelImg)
    valImgList.append(cp.asarray(valImg))
    vallabelList.append(cp.asarray(valLabelImg))
    

tf.compat.v1.logging.set_verbosity(tf.compat.v1.logging.ERROR)
workbook = openpyxl.Workbook()
modelIndex = 3
#for model_path in [model_path_sphereRandom, model_path_sphere, model_path_Random, model_path_Conventional]:   
for model_path in [model_path_Conventional]:    
    excel = [["caseID", "DSC"]]
    model_List = os.listdir(model_path)
    model_List.sort()
    model_List.reverse()
    count = 1
    for md in range(len(model_List)): 
        dice = []
            
        sigma = int(model_List[md].replace(".h5","").split("sigma_")[1].split("_")[0])  
        beta = int(model_List[md].replace(".h5","").split("beta_")[1].split("_")[0])
        
        if modelIndex == 3:
            valProcessed, valLabelProcessed = preprocessing_val(valImgList, vallabelList, sigma, beta, mean, dev)
        else:
            valProcessed, valLabelProcessed = preprocessing_pro(valImgList, vallabelList, sigma, beta, mean, dev)
        
        model = load_model(model_path + model_List[md])
        for x in range(len(valProcessed)):  
            p = tf.cast(model.predict({'main_input': cp.asnumpy((valProcessed[x])[cp.newaxis, :, :, :, :])}), tf.float64)      
            p_r = np.where(p>=0.5, 1.0, 0.0)
            dice.append(dice_coef(cp.asnumpy(valLabelProcessed[x]), p_r[0, :, :, :, :]))
        excel.append([model_List[md], np.mean(dice)])        
        print(str(count) + "/" + str(len(model_List)), model_List[md] + '\t', float(np.mean(dice)))
        count = count + 1
    index = len(excel)
    sheet = workbook.create_sheet(title=model_path.split('//')[-2], index = 0)
    modelIndex = modelIndex + 1
    for i in range(0, index):
        for j in range(0, len(excel[i])):
            sheet.cell(row=i+1, column=j+1, value=excel[i][j])
workbook.save('kernelResult_test_conventional.xlsx')
'''
    excel = [["caseID", "DSC of without morphology"]]
    
    for x in range(len(dicePre)):
        excel.append([valImgName[x], dicePre[x]])
            
    for weights in range(10):
        diceFinal = []
        index = 0
        excel[0].append("DSC of with morphology, w = %d" %(weights))
        for x in morphoDataList:
            temp = cp.zeros((40, 40, 40))
            factor = 0
            for img in x:
                p = model1.predict({'main_input': img[cp.newaxis, :, :, :, :]})
                r = p[0, :, :, :, 0]
                temp = temp + r
                factor = factor + 1
            result = (cp.round((temp + predPre[index] * weights) / (factor + weights))).astype(cp.float32)
            #result = cp.where((temp / factor) >= 2/3, 1, 0).astype(cp.float32)
            diceFinal.append(dice_coef(valLabelProcessed[index][:, :, :, 0], result, smooth = 0.000001).numpy())
            predFinal.append(result[cp.newaxis, :, :, :, cp.newaxis])
            index = index + 1
            
        for x in range(1, len(excel)):
            excel[x].append(diceFinal[x - 1])
        
    write_excel_xlsx("F:\\Cui\\RL+DVN\\TeacherTraining\\FinalResult.xlsx", 'result', excel)
'''
'''
    index = 0
    for x in morphoDataList:
        temp = cp.zeros((40, 40, 40))
        factor = 0
        for img in x:
            p = model1.predict({'main_input': img[cp.newaxis, :, :, :, :]})
            r = p[0, :, :, :, 0]
            temp = temp + r
            factor = factor + 1
        result = (cp.round((temp + predPre[index] * factor) / (factor + factor))).astype(cp.float32)
        diceFinal.append(dice_coef(valLabelProcessed[index][:, :, :, 0], result, smooth = 0.000001).numpy())
        predFinal.append(result[cp.newaxis, :, : , :, cp.newaxis])
        #predFinal.append(cp.where(temp >= factor*2//3, 1, 0))
        index = index + 1
        
    excel = [["caseID", "DSC of without morphology", "DSC of with morphology"]]
    
    for x in range(len(dicePre)):
        excel.append([valImgName[x], dicePre[x], diceFinal[x]])
        
    write_excel_xlsx("F:\\Cui\\RL+DVN\\TeacherTraining\\FinalResult.xlsx", 'result', excel)
'''
    
        
'''
pred = predFinal
for im in range(len(pred)):
    kernel = generate_mask(7, 7, 7, 3, 3, 3, 3)
    #kernel = generate_mask(5, 5, 5, 2, 2, 2, 2)
    #kernel = generate_mask(3, 3, 3, 1, 1, 1, 1)
    hm, d, e = morphologyHintMap(cp.where(pred[im]>0.5, 1, 0), kernel = kernel)
    HM.append(hm)
    dliList.append(d)
    eroList.append(e)
    cv2.imwrite(savePNG + valImgName[im] + '_hintmap.png', HM[im][:, :, HM[im].shape[-1] // 2] * 255)
    cv2.imwrite(savePNG + valImgName[im] + '_pred.png', pred[im][0, :, :, HM[im].shape[-1] // 2, 0] * 255)
    hm, d, e = morphologyHintMap(valLabelProcessed[im][:, :, :, 0], kernel = kernel)
    HMLabel.append(hm)
    dliListLabel.append(d)
    eroListLabel.append(e)
    cv2.imwrite(savePNG + valImgName[im] + '_hintmap_label.png', HMLabel[im][:, :, HM[im].shape[-1] // 2] * 255)
    cv2.imwrite(savePNG + valImgName[im] + '_label.png', valLabelProcessed[im][:, :, HM[im].shape[-1] // 2, 0] * 255)
    cv2.imwrite(savePNG + valImgName[im] + '_dli_label.png', dliListLabel[im][:, :, 20] * 255)
    cv2.imwrite(savePNG + valImgName[im] + '_ero_label.png', eroListLabel[im][:, :, 20] * 255)
    
meanTumorB, meanBack = toGetBand.FillMean()
meanBackAvg = meanBack

for x in range(len(dliList)):
    reversedDliList.append(1 - dliList[x])
    reversedDliLabelList.append(1 - dliListLabel[x])
    meanBack = min(cp.min(valProcessed[x]), meanBack)*56.15427+71.85165
    m = cp.mean([valProcessed[x][HM[x].shape[0] // 2, HM[x].shape[1] // 2, HM[2].shape[-1] // 2, 0], \
                valProcessed[x][-1 + HM[x].shape[0] // 2, HM[x].shape[1] // 2, HM[2].shape[-1] // 2, 0], \
                valProcessed[x][-1 + HM[x].shape[0] // 2, -1 + HM[x].shape[1] // 2, HM[2].shape[-1] // 2, 0], \
                valProcessed[x][-1 + HM[x].shape[0] // 2, -1 + HM[x].shape[1] // 2, -1 + HM[2].shape[-1] // 2, 0], \
                valProcessed[x][-1 + HM[x].shape[0] // 2, HM[x].shape[1] // 2, -1 + HM[2].shape[-1] // 2, 0], \
                valProcessed[x][HM[x].shape[0] // 2, -1 + HM[x].shape[1] // 2, HM[2].shape[-1] // 2, 0], \
                valProcessed[x][HM[x].shape[0] // 2, -1 + HM[x].shape[1] // 2, -1 + HM[2].shape[-1] // 2, 0], \
                valProcessed[x][HM[x].shape[0] // 2, HM[x].shape[1] // 2, -1 + HM[2].shape[-1] // 2, 0]])
    meanTumor = max(m, meanTumorB)*56.15427+71.85165
    revised = HM[x] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0] + reversedDliList[x] * meanBack + eroList[x] * meanTumor * 0.5 + eroList[x] * 0.5 * (valProcessed[x] * 56.15427+71.85165)[:, :, :, 0]
    revisedlabel = HMLabel[x] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0] + reversedDliLabelList[x] * meanBack + eroListLabel[x] * meanTumor * 0.5 + eroListLabel[x]  * 0.5 * (valProcessed[x] * 56.15427+71.85165)[:, :, :, 0]
    
    revised[revised < 0] = 0
    revisedlabel[revisedlabel < 0] = 0
    revised[revised > 255] = 255
    revisedlabel[revisedlabel > 255] = 255

    notRevised = HM[x][:, :, :] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0]
    notRevisedLabel = HMLabel[x][:, :, :] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0]
    cv2.imwrite(savePNG + valImgName[x] + '_revised.png', revised[:, :, revised.shape[2] // 2])
    cv2.imwrite(savePNG + valImgName[x] + '_revised_label.png', revisedlabel[:, :, revisedlabel.shape[2] // 2])
    t = Edge_Extract((revised).astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_revised&ref&pred.png', t[:, :, revised.shape[2] // 2])
    t = Edge_Extract((revisedlabel).astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_revised&ref&pred_label.png', t[:, :, revised.shape[2] // 2])
    t = Edge_Extract_onlyRef((revised).astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_revised&ref.png', t[:, :, t.shape[-2] // 2, :])
    t = Edge_Extract_onlyPred((revised).astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_revised&pred.png', t[:, :, t.shape[-2] // 2, :])
    nt = Edge_Extract_onlyRef((notRevised).astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_notRevised&ref.png', nt[:, :, nt.shape[-2] // 2, :])
    
    t = Edge_Extrac_final((revised).astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_preFinald&ref&revised.png', t[:, :, t.shape[-2] // 2, :])  
    t = Edge_Extract_onlyRef((revisedlabel).astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    nt = Edge_Extract_onlyRef((notRevisedLabel).astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_revised&ref_label.png', t[:, :, t.shape[-2] // 2, :])
    cv2.imwrite(savePNG + valImgName[x] + '_notRevised&ref_label.png', nt[:, :, nt.shape[-2] // 2, :])
    #Image.fromarray((HM[x][:, :, HM[x].shape[-1] // 2] * (valProcessed[x]*56.15427+71.85165)[:, :, HM[x].shape[-1] // 2, 0] + reversedDliList[x][:, :, HM[x].shape[-1] // 2] * meanBack + eroList [x][:, :, HM[x].shape[-1] // 2] * meanTumor).astype(cp.uint8)).save(savePNG + valImgName[x] + '_revised.png')
    #imageio.imsave(savePNG + valImgName[x] + '_revised.png', (HM[x][:, :, HM[x].shape[-1] // 2] * (valProcessed[x]*56.15427+71.85165)[:, :, HM[x].shape[-1] // 2, 0] + reversedDliList[x][:, :, HM[x].shape[-1] // 2] * meanBack + eroList [x][:, :, HM[x].shape[-1] // 2] * meanTumor).astype(cp.uint8))

for x in range(len(pred)):
    dice.append(dice_coef(valLabelProcessed[x], pred[x], smooth = 0.000001).numpy())
    haus.append(computeQualityMeasures(valLabelProcessed[x][:, :, :, 0].astype(cp.uint8), pred[x][0, :, :, :, 0].astype(cp.uint8)))
    img8bit = valProcessed[x]*56.15427+71.85165
    img8bit[img8bit < 0] = 0
    img8bit[img8bit > 255] = 255
    cv2.imwrite(savePNG + valImgName[x] + '_origin.png', img8bit[:, :, HM[x].shape[-1] // 2, 0])
    cv2.imwrite(savePNG + valImgName[x] + '_masked.png', HM[x][:, :, HM[x].shape[-1] // 2] * img8bit[:, :, HM[x].shape[-1] // 2, 0])
    cv2.imwrite(savePNG + valImgName[x] + '_masked_label.png', HMLabel[x][:, :, HMLabel[x].shape[-1] // 2] * img8bit[:, :, HMLabel[x].shape[-1] // 2, 0])
    t = Edge_ExtractFor3((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    edged.append(t.astype(cp.uint8))
    edged3D = toList(t.astype(cp.uint8))
    edged3D[0].save(saveGIF + valImgName[x].split(".nii")[0] + '.gif',
           save_all=True, append_images=edged3D[1:], loop = 0, duration = 200)
    hintmap = Gray2RGB3D(HM[x])
    fourcc = cv2.VideoWriter_fourcc('m', 'p', '4', 'v')
    out = cv2.VideoWriter(saveVideo + valImgName[x].split(".nii")[0] + '.mp4',fourcc, 5.0, (40,40))
    for y in range(t.shape[-2]):
        out.write(t.astype(cp.uint8)[:, :, y, :])
    out.release()
    cv2.destroyAllWindows()
    t = Edge_Extract((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_masked&edge.png', hintmap[:, :, hintmap.shape[-2] // 2, :] * t[:, :, t.shape[-2] // 2, :])
    cv2.imwrite(savePNG + valImgName[x] + '_pred&ref.png', t[:, :, t.shape[-2] // 2, :])    
    cv2.imwrite(savePNG + valImgName[x] + '_dli.png', dliList[x][:, :, t.shape[-2] // 2] * 255)
    cv2.imwrite(savePNG + valImgName[x] + '_ero.png', eroList[x][:, :, t.shape[-2] // 2] * 255)
    r = Edge_Extract_onlyRef((img8bit).astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_ref.png', r[:, :, t.shape[-2] // 2, :])
    t = Edge_Extract_onlyPred((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_pred.png', t[:, :, t.shape[-2] // 2, :])
    t = Edge_ExtractFor3((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_pred&predFinal.png', t[:, :, t.shape[-2] // 2, :])
    t = Edge_Extract_onlyPredFinal((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_predFinal.png', t[:, :, t.shape[-2] // 2, :])
    t = Edge_Extract_Ref(255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
    cv2.imwrite(savePNG + valImgName[x] + '_refonly.png', t[:, :, t.shape[-2] // 2, :])

for x in range(len(pred)):   
     for slices in range(-2, 6, 1):
         img8bit = valProcessed[x]*56.15427+71.85165
         notRevised = HM[x][:, :, :] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0]
         revised = HM[x] * (valProcessed[x]*56.15427+71.85165)[:, :, :, 0] + reversedDliList[x] * meanBack + eroList[x] * meanTumor * 0.5 + eroList[x] * 0.5 * (valProcessed[x] * 56.15427+71.85165)[:, :, :, 0]
         img8bit[img8bit < 0] = 0
         img8bit[img8bit > 255] = 255
         revised[revised < 0] = 0
         revised[revised > 255] = 255
         notRevised[notRevised < 0] = 0
         notRevised[notRevised > 255] = 255
         cv2.imwrite(savePNG_slices + valImgName[x] + '_slices_' + str(slices) + '_origin.png', img8bit[:, :, slices + HM[x].shape[-1] // 2, 0])
         t = Edge_Extract_onlyPred((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8))
         t[t < 0] = 0
         t[t > 255] = 255
         cv2.imwrite(savePNG_slices + valImgName[x] + '_slices_' + str(slices) + '_pred.png', t[:, :, slices + t.shape[-2] // 2, :])
         cv2.imwrite(savePNG_slices + valImgName[x] + '_slices_' + str(slices) + '_revised.png', revised[:, :, slices + revised.shape[2] // 2])
         cv2.imwrite(savePNG_slices + valImgName[x] + '_slices_' + str(slices) + '_notRevised.png', notRevised[:, :, slices + revised.shape[2] // 2])
         t = Edge_ExtractFor3((img8bit).astype(cp.uint8), 255 * predPre[x].astype(cp.uint8), 255 * pred[x][0, :, :, :, 0].astype(cp.uint8), 255 * valLabelProcessed[x][:, :, :, 0].astype(cp.uint8))
         t[t < 0] = 0
         t[t > 255] = 255
         cv2.imwrite(savePNG_slices + valImgName[x] + '_slices_' + str(slices) + '_pred&predFinal&ref.png', t[:, :, slices + t.shape[-2] // 2, :])
'''